"""getting info from a spreasheet and then
asking the user to type more info
to fill several spreadsheets in a specific format
"""
import datetime
from os import listdir
from openpyxl import Workbook
from openpyxl import load_workbook
now = datetime.datetime.now()
MES = now.strftime('%m')
ANO = now.strftime('%Y')
FILES = [f for f in listdir() if (('.xlsx' in f) and (f'{ANO}{MES}' in f))][0]
workbook = Workbook()
workbook = load_workbook(
    filename=FILES)
MOV_PLANILHA = {}
LINHAS = 2
VALOR = 0
for i in workbook.sheetnames:
    sheet = workbook[i]
    if i == 'ProgramacaoTrecho':
        while sheet[f'A{LINHAS}'].value == (f'2024-{MES}'):
            LINHAS += 1
        for j in range(2, LINHAS):
            IGNORE = 'no'
            try:
                if ("D11" in str(sheet[f'G{j}'].value)) or \
                    ("D12" in str(sheet[f'G{j}'].value)) or \
                    ("D13" in str(sheet[f'G{j}'].value)) or \
                        ("D14" in str(sheet[f'G{j}'].value)):
                    IGNORE = 'yes'
                elif ("D21" in str(sheet[f'G{j}'].value)) or \
                    ("D22" in str(sheet[f'G{j}'].value)) or \
                    ("D23" in str(sheet[f'G{j}'].value)) or \
                        ("D24" in str(sheet[f'G{j}'].value)):
                    IGNORE = 'yes'
                elif str(sheet[f'O{j}'].value) == "RE":
                    IGNORE = 'yes'
                if IGNORE != "yes":
                    VALOR = MOV_PLANILHA[str(
                        sheet[f'D{j}'].value)+' '+str(sheet[f'G{j}'].value)]
                    if ((str(sheet[f'G{j}'].value)) == 'D3') or \
                            ((str(sheet[f'G{j}'].value)) == 'D4'):
                        if VALOR > sheet[f'K{j}'].value/1000:
                            pass
                        else:
                            MOV_PLANILHA[str(
                                sheet[f'D{j}'].value)+' '+str(sheet[f'G{j}'].value)] = sheet[
                                    f'K{j}'].value/1000
                    elif ((str(sheet[f'G{j}'].value)) == 'D5') and \
                            ((str(sheet[f'E{j}'].value)) == 'P1'):
                        if VALOR > sheet[f'K{j}'].value/1000:
                            MOV_PLANILHA[str(
                                sheet[f'D{j}'].value)+' '+str(sheet[
                                    f'G{j}'].value)+' (IN)'] = sheet[f'K{j}'].value/1000
                        else:
                            MOV_PLANILHA[str(
                                sheet[f'D{j}'].value)+' '+str(sheet[f'G{j}'].value)] = sheet[
                                    f'K{j}'].value/1000
                            MOV_PLANILHA[str(
                                sheet[f'D{j}'].value)+' '+str(sheet[f'G{j}'].value)+' (IN)'] = VALOR

                    else:
                        MOV_PLANILHA[str(
                            sheet[f'D{j}'].value)+' '+str(sheet[f'G{j}'].value)] = VALOR+sheet[
                                f'K{j}'].value/1000
            except Exception:
                if IGNORE != "yes":
                    MOV_PLANILHA[str(
                        sheet[f'D{j}'].value)+' '+str(sheet[f'G{j}'].value)] = sheet[
                            f'K{j}'].value/1000
print('Valores de Movimentação obtidos')
print('\nMovimentações de Duto')
oc = {'D6': [], 'D7': [], 'D8': [],
      'D9': [], 'P12': [], 'D5': [], 'D11': []}
for i in oc:
    try:
        P1 = MOV_PLANILHA['P1 '+i]
        P2 = MOV_PLANILHA['P2 '+i]
    except KeyError:
        print(i)
        P1 = float(input('Digite P1: '))
        P2 = float(input('Digite P2: '))
    oc[i] = [P1, P2]
craq = {'D5': []}
for i in craq:
    try:
        P1 = MOV_PLANILHA['P1 '+i+' (IN)']
    except KeyError:
        print(i)
        P1 = float(input('Digite P1: '))
    craq[i] = [P1]
print('ok')

print('\nP3')
g = {'D12': [], 'D13': [], 'D14': [], 'D15': []}
for i in g:
    try:
        if i == "D15":
            P5 = 0
            P4 = 0
            P3 = MOV_PLANILHA['P3 '+i+' I']
        elif i == 'D12':
            P5 = 0
            P4 = MOV_PLANILHA['P4 '+i]
            P3 = MOV_PLANILHA['P3 '+i]
        elif i == 'D14':
            P5 = MOV_PLANILHA['P3 '+i]
            P4 = 0
            P3 = 0
        else:
            P5 = MOV_PLANILHA['P3 '+i]
            P4 = MOV_PLANILHA['P4 '+i]
            P3 = MOV_PLANILHA['P3 '+i]
    except KeyError:
        print(i)
        P5 = float(input('Digite P5:'))
        P4 = float(input('Digite P4:'))
        P3 = float(input('Digite P3:'))
    g[i] = [P5, P4, P3]
print('ok')

nft = {'D16': [], 'D4': [], 'D3': [],
       'D17': [], 'D18-III': [], 'D19': [],
       'D20': [0], 'D21': [0]}
for i in nft:
    try:
        if i == 'D19':
            P6 = MOV_PLANILHA['P6 '+'D19']
        else:
            P6 = MOV_PLANILHA['P6 '+i]
    except KeyError:
        print(i)
        P6 = float(input('Digite P6:'))
    nft[i] = [P6]
print('ok')
glna = {'D22': [], 'D1': [], 'D4': [], 'D21': [], 'D16': [], 'D2': [], 'D3': [
], 'D17': [], 'D23': [], 'D18-III': [], 'D24': [], 'D25': [], 'D26': [], 'D27': [],
    'D20': [0, 0, 0, 0]}
for i in glna:
    try:
        if i in ("D22", 'D1', 'D4', 'D21', 'D23', 'D24', 'D25'):
            P7 = MOV_PLANILHA['P7 '+i]
            P8 = 0
            P9 = 0
            P10 = P7+P8
        elif i in ('D16', 'D2', 'D3', 'D17', 'D26', 'D27'):
            P7 = MOV_PLANILHA['P7 '+i]
            P8 = MOV_PLANILHA['P8 '+i]
            P9 = 0
            P10 = P7+P8
        elif i == 'D18-III':
            P7 = 0
            P8 = MOV_PLANILHA['P8 '+i]
            P9 = 0
            P10 = 0
    except KeyError:
        print(i)
        P7 = float(input('Digite P7:'))
        P8 = float(input('Digite P8:'))
        P9 = float(input('Digite P9:'))
        P10 = float(input('Digite P7 somada:'))
    glna[i] = [P7, P8, P9, P10]  # pylint: disable=E0606
print('ok')
dsl = {'D28': [], 'D1': [], 'D29': [], 'D20': [], 'D16': [], 'D2': [], 'D3': [
], 'D30': [], 'D31': [], 'D32': [], 'D33': [], 'D34': [], 'D25': [], 'D35': [], 'D36': [
], 'D37': [], 'D38': [], 'D39': [], 'D4': [0, 0, 0, 0], 'D21': [0, 0, 0, 0]}
for i in dsl:
    try:
        if i in ('D28', 'D31', 'D32'):
            P11 = MOV_PLANILHA['P11 '+i]
            P12 = 0
            P13 = 0
            P14 = 0
        elif i == 'D1':
            P11 = MOV_PLANILHA['P11 '+i]
            P12 = MOV_PLANILHA['P12 '+i]
            P13 = 0
            P14 = MOV_PLANILHA['P14 '+i]
        elif i in ('D29', 'D33', 'D34', 'D25', 'D35', 'D36', 'D38'):
            P11 = 0
            P12 = MOV_PLANILHA['P12 '+i]
            P13 = 0
            P14 = 0
        elif i in ('D20', 'D30'):
            P11 = MOV_PLANILHA['P11 '+i]
            P12 = MOV_PLANILHA['P12 '+i]
            P13 = 0
            P14 = 0
        elif i in ('D2', 'D37', 'D39', 'D16'):
            P11 = MOV_PLANILHA['P11 '+i]
            P12 = 0
            P13 = MOV_PLANILHA['P13 '+i]
            P14 = 0
        elif i == 'D3':
            P11 = 0
            P12 = 0
            P13 = MOV_PLANILHA['P13 '+i]
            P14 = 0
    except KeyError:
        print(i)
        P11 = float(input('Digite P11:'))
        P12 = float(input('Digite P12:'))
        P13 = float(input('Digite P13:'))
        P14 = float(input('Digite P14:'))
    dsl[i] = [P11, P12, P13, P14]  # pylint: disable=E0606
print('ok')

qa = {'D4': [], 'D30': [], 'D3': [], 'D31': [], 'D32': [],
      'D20': [0], 'D16': [0], 'D21': [0]}
for i in qa:
    try:
        P15 = MOV_PLANILHA['P15 '+i]
    except KeyError:
        print(i)
        P15 = float(input('Digite P15:'))
    qa[i] = [P15]
print('ok')

eol = {'TR_D22': [], 'TR_D4': [], 'D21': [], 'D2': [], 'D23': [0, 0]}
for i in eol:
    try:
        if i == "TR_D22":
            P16 = MOV_PLANILHA['P16 '+'D22']
            P17 = MOV_PLANILHA['P17 '+'D22']
        elif i == "TR_D4":
            P16 = MOV_PLANILHA['P16 '+'D4']
            P17 = MOV_PLANILHA['P17 '+'D4']
        else:
            P16 = MOV_PLANILHA['P16 '+i]
            P17 = MOV_PLANILHA['P17 '+i]
    except KeyError:
        print(i)
        P16 = float(input('Digite P16:'))
        P17 = float(input('Digite P17:'))
    eol[i] = [P16, P17]
print('ok')

PO1 = {'P6': [], 'P7': [], 'P8': [], 'P15': [],
       'P12': [], 'P11': [], 'P13': [], 'P1': [], 'P2': []}
PO2 = {'P3': [], 'P7': [], 'P8': [], 'P12': [],
       'P11': [], 'P13': [], 'P2': [], 'P1': [], 'P5': []}
for i in PO1:
    print(i)
    while True:
        try:
            carga_PO1 = float(input('Digite Carga e/ou Descarga:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    PO1[i] = [carga_PO1]
print('\nCarga e Descarga - PO2')
for i in PO2:
    print(i)
    while True:
        try:
            carga_PO2 = float(input('Digite Carga e/ou Descarga:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    PO2[i] = [carga_PO2]

print('\nPLANO DE PRODUÇÃO - R1')
R1 = {'P1': [], 'P2': [], 'P6': [], 'P7': [], 'P8': [],
      'P15': [], 'P12': [], 'P11': [], 'P13': [], 'P14': [], 'P3': []}
for i in R1:
    print(i)
    while True:
        try:
            prod_R1 = float(input('Digite Plano de Produção:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    R1[i] = [prod_R1]
print('\nPLANO DE PRODUÇÃO - R2')
R2 = {'P3': [], 'P6': [], 'P7': [], 'P8': [], 'P12': [],
      'P11': [], 'P13': [], 'P2': [], 'P1': [], 'P15': [], 'P18': []}
for i in R2:
    print(i)
    while True:
        try:
            prod_R2 = float(input('Digite Plano de Produção:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    R2[i] = [prod_R2]
print('\nPLANO DE PRODUÇÃO - R3')
R3 = {'P3': [], 'P6': [], 'P7': [],
      'PPBR': [], 'P12': [], 'P2': [], 'P1': []}
for i in R3:
    print(i)
    while True:
        try:
            prod_R3 = float(input('Digite Plano de Produção:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    R3[i] = [prod_R3]
print('\nPLANO DE PRODUÇÃO - R4')
R4 = {'PRIN': [], 'P3': [], 'P7': [], 'P8': [],
      'P12': [], 'P11': [], 'P13': [], 'P2': [], 'P1': []}
for i in R4:
    print(i)
    while True:
        try:
            prod_R4 = float(input('Digite Plano de Produção:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    R4[i] = [prod_R4]
print('\nPLANO DE PRODUÇÃO - R5')
R5 = {'P3': [], 'P7': [], 'P12': [], 'P11': [], 'P13': []}
for i in R5:
    print(i)
    while True:
        try:
            if i == 'P3':
                prod_R5 = float(input('Digite Plano de Produção (ton):'))/0.55
            else:
                prod_R5 = float(input('Digite Plano de Produção:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    R5[i] = [prod_R5]

print('\nMERCADO - ESCUROS')
m_escuros = {'P1 TA1': [], 'P2 TA1': [],
             'P1 PO1': [], 'P2 PO1': []}
for i in m_escuros:
    print(i)
    while True:
        try:
            merc_ESCUROS = float(input('Digite Mercado:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    m_escuros[i] = [merc_ESCUROS]

print('\nMERCADO - P3')
m_P3 = {'R1': [], 'R2': [], 'R3': [], 'TA1': []}
for i in m_P3:
    print(i)
    while True:
        try:
            merc_P3 = float(input('Digite Mercado:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    m_P3[i] = [merc_P3]

print('\nMERCADO - JETA')
m_jet = {'R1': [], 'T1': []}
for i in m_jet:
    print(i)
    while True:
        try:
            merc_JET = float(input('Digite Mercado:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    m_jet[i] = [merc_JET]

print('\nMERCADO - P6')
m_naf = {'100%': []}
for i in m_naf:
    print(i)
    while True:
        try:
            merc_NFT = float(input('Digite Mercado:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    m_naf[i] = [merc_NFT]
m_naf['95%'] = [m_naf['100%'][0]*0.95]

print('\nMERCADO - REFINARIAS')
m_ref = {'R3': [], 'R1': [], 'R2': [], 'R4': []}
for i in m_ref:
    print(i)
    while True:
        try:
            merc_REF_GLN = float(input('Digite Mercado de P7:'))
            merc_REF_P15 = float(input('Digite Mercado de P15:'))
            merc_REF_P13 = float(input('Digite Mercado de P13:'))
            merc_REF_S500 = float(input('Digite Mercado de P11:'))
            merc_REF_S10 = float(input('Digite Mercado de P12:'))
            merc_REF_PPBR = float(input('Digite Mercado de PPBR:'))
            merc_REF_P3 = float(input('Digite Mercado de P3:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    m_ref[i] = [merc_REF_GLN, merc_REF_P15, merc_REF_P13, merc_REF_S500, merc_REF_S10,
                merc_REF_PPBR, merc_REF_P3]
print('\nMERCADO - OPASC (BANDEIRA BR)')
m_opasc = {'R5': [], 'T2': [],
           'T3': [], 'T4': []}
for i in m_opasc:
    print(i)
    while True:
        try:
            if i in ('R5', 'T3'):
                merc_OPASC_P3 = float(input('Digite Mercado de P3(ton):'))/0.55
                merc_OPASC_P13 = float(input('Digite Mercado de P13:'))
            else:
                merc_OPASC_P3 = 0
                merc_OPASC_P13 = 0
            merc_OPASC_GLNA = float(input('Digite Mercado de P7:'))
            merc_OPASC_S10 = float(input('Digite Mercado de P12:'))
            merc_OPASC_S500 = float(input('Digite Mercado de P11:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    m_opasc[i] = [merc_OPASC_P3, merc_OPASC_GLNA,
                  merc_OPASC_S10, merc_OPASC_S500, merc_OPASC_P13]

print('\nMERCADO - TERMINAIS')
m_terminais = {'T5': [], 'T6': [], 'T7': [], 'T8': [
], 'T9': [], 'T10': [], 'T11': [], 'T12': [], 'TA TA1': [
], 'T13': [], 'T14': [], 'T15': []}
for i in m_terminais:
    print(i)
    while True:
        try:
            if i in ('T10', 'TA TA1'):
                m_terminais_P1 = float(
                    input('Digite Mercado de P1(ton):'))/0.93
            else:
                m_terminais_P1 = 0
            if i == 'TA TA1':
                m_terminais_P5 = float(input('Digite Mercado de P5:'))
                m_terminais_P3 = float(
                    input('Digite Mercado de P3(ton):'))/0.55
                m_terminais_P13 = float(input('Digite Mercado de P13:'))
            else:
                m_terminais_P5 = 0
                m_terminais_P3 = 0
                m_terminais_P13 = 0
            if i == 'T13':
                m_terminais_P15 = float(input('Digite Mercado de P15:'))
            else:
                m_terminais_P15 = 0
            m_terminais_GLNA = float(input('Digite Mercado de P7:'))
            m_terminais_S500 = float(input('Digite Mercado de P11:'))
            m_terminais_S10 = float(input('Digite Mercado de P12:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    m_terminais[i] = [m_terminais_GLNA, m_terminais_P15, m_terminais_P13,
                      m_terminais_S500, m_terminais_S10, m_terminais_P5,
                      m_terminais_P3, m_terminais_P1]

emeta_refinarias = {'R3': [], 'R1': [], 'R2': [], 'R4': []}
for k in emeta_refinarias:
    print(k)
    while True:
        try:
            emeta_refinarias_NFT = float(input('Digite Estoque Meta de P6:'))
            emeta_refinarias_GLNA = float(input('Digite Estoque Meta de P7:'))
            emeta_refinarias_P8 = float(input('Digite Estoque Meta de P8:'))
            emeta_refinarias_P15J = float(input('Digite Estoque Meta de P15:'))
            emeta_refinarias_P13 = float(input('Digite Estoque Meta de P13:'))
            emeta_refinarias_S500 = float(input('Digite Estoque Meta de P11:'))
            emeta_refinarias_S10 = float(input('Digite Estoque Meta de P12:'))
            emeta_refinarias_P3 = float(
                input('Digite Estoque Meta de P3(ton):'))/0.55
            emeta_refinarias_P1 = float(
                input('Digite Estoque Meta de P1(ton):'))/0.93
            emeta_refinarias_P2 = float(
                input('Digite Estoque Meta de P2(ton):'))/0.93
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    emeta_refinarias[k] = [emeta_refinarias_NFT, emeta_refinarias_GLNA,
                           emeta_refinarias_P8, emeta_refinarias_P15J,
                           emeta_refinarias_P13, emeta_refinarias_S500,
                           emeta_refinarias_S10, emeta_refinarias_P3,
                           emeta_refinarias_P1, emeta_refinarias_P2]

emeta_terminais = {'T5': [], 'T6': [], 'T7': [], 'T8': [
], 'T9': [], 'T10': [], 'T11': [], 'T12': [], 'TA TA1': [
], 'T13': [], 'T14': [], 'T15': []}
for i in emeta_terminais:
    print(i)
    while True:
        try:
            if i in ('T10', 'TA TA1', 'T14', 'T15'):
                emeta_terminais_P1 = float(
                    input('Digite Estoque Meta de P1(ton):'))/0.93
                emeta_terminais_P13 = float(
                    input('Digite Estoque Meta de P13:'))
            else:
                emeta_terminais_P1 = 0
                emeta_terminais_P13 = 0
            if i == 'TA TA1':
                emeta_terminais_P3 = float(
                    input('Digite Estoque Meta de P3(ton):'))/0.55
            else:
                emeta_terminais_P3 = 0
            if i == 'T10':
                emeta_terminais_NFT = float(
                    input('Digite Estoque Meta de P6:'))
            else:
                emeta_terminais_NFT = 0
            if i == 'T13' or i == 'T11' or i == 'T10':
                emeta_terminais_P15 = float(
                    input('Digite Estoque Meta de P15:'))
            else:
                emeta_terminais_P15 = 0
            emeta_terminais_GLNA = float(input('Digite Estoque Meta de P7:'))
            emeta_terminais_S500 = float(input('Digite Estoque Meta de P11:'))
            emeta_terminais_S10 = float(input('Digite Estoque Meta de P12:'))
        except NameError:
            print('Digite de novo')
            continue
        else:
            break
    emeta_terminais[i] = [emeta_terminais_NFT, emeta_terminais_GLNA,
                          emeta_terminais_P15, emeta_terminais_P13,
                          emeta_terminais_S500, emeta_terminais_S10,
                          emeta_terminais_P3, emeta_terminais_P1]
workbook = Workbook()
workbook = load_workbook(filename=f"bandeira{ANO}{MES}.xlsx")
for i in workbook.sheetnames:
    IND = 1
    sheet = workbook[i]
    if i == "Bandeira-Painel":
        while IND < 98:
            if sheet[f"B{IND}"].value in ('P7', 'GLN', 'P9', 'P7+P8', 'P8'):
                if sheet[f"B{IND}"].value == 'P7':
                    sheet[f"C{IND}"] = glna[sheet[f"A{IND}"].value][0]
                elif sheet[f"B{IND}"].value == 'P8':
                    sheet[f"C{IND}"] = glna[sheet[f"A{IND}"].value][1]
                elif sheet[f"B{IND}"].value == 'P9':
                    sheet[f"C{IND}"] = glna[sheet[f"A{IND}"].value][2]
                else:
                    sheet[f"C{IND}"] = glna[sheet[f"A{IND}"].value][3]

            elif sheet[f"B{IND}"].value == 'P6':
                sheet[f"C{IND}"] = P6[sheet[f"A{IND}"].value][0]
            elif sheet[f"B{IND}"].value == 'P15':
                sheet[f"C{IND}"] = qa[sheet[f"A{IND}"].value][0]

            elif sheet[f"B{IND}"].value in ('P13', 'P12', 'P11', 'S501', 'P14'):
                if sheet[f"B{IND}"].value in ('P11', 'S501'):
                    sheet[f"C{IND}"] = dsl[sheet[f"A{IND}"].value][0]
                elif sheet[f"B{IND}"].value == 'P12':
                    sheet[f"C{IND}"] = dsl[sheet[f"A{IND}"].value][1]
                elif sheet[f"B{IND}"].value == 'P13':
                    sheet[f"C{IND}"] = dsl[sheet[f"A{IND}"].value][2]
                else:
                    sheet[f"C{IND}"] = dsl[sheet[f"A{IND}"].value][3]

            elif sheet[f"B{IND}"].value in ('P16', 'P17'):
                if sheet[f"B{IND}"].value == 'P16':
                    sheet[f"C{IND}"] = eol[sheet[f"A{IND}"].value][0]
                else:
                    sheet[f"C{IND}"] = eol[sheet[f"A{IND}"].value][1]

            elif sheet[f"B{IND}"].value in ('PRIN', 'P3', 'P5'):
                if sheet[f"B{IND}"].value == 'PRIN':
                    sheet[f"C{IND}"] = g[sheet[f"A{IND}"].value][1]
                elif sheet[f"B{IND}"].value == 'P5':
                    sheet[f"C{IND}"] = g[sheet[f"A{IND}"].value][0]
                else:
                    sheet[f"C{IND}"] = g[sheet[f"A{IND}"].value][2]
            elif sheet[f"B{IND}"].value in ('P1', 'P2'):
                if sheet[f"B{IND}"].value == 'P1':
                    sheet[f"C{IND}"] = oc[sheet[f"A{IND}"].value][0]
                else:
                    sheet[f"C{IND}"] = oc[sheet[f"A{IND}"].value][1]
            IND += 1
    elif i == "Semanal":
        IND = 0
        indice = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                  'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'W', 'Y', 'Z', 'AA',
                  'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM',
                  'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AW', 'AY',
                  'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL',
                  'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BW',
                  'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK',
                  'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU']
        while IND < 102:
            if sheet[f"{indice[IND]}2"].value in ('P7', 'GLN', 'P9', 'P7+P8', 'P8'):
                if sheet[f"{indice[IND]}2"].value == 'P7':
                    sheet[f"{indice[IND]}3"] = glna[sheet[f"{
                        indice[IND]}1"].value][0]
                elif sheet[f"{indice[IND]}2"].value == 'P8':
                    sheet[f"{indice[IND]}3"] = glna[sheet[f"{
                        indice[IND]}1"].value][1]
                elif sheet[f"{indice[IND]}2"].value == 'P9':
                    sheet[f"{indice[IND]}3"] = glna[sheet[f"{
                        indice[IND]}1"].value][2]
                else:
                    sheet[f"{indice[IND]}3"] = glna[sheet[f"{
                        indice[IND]}1"].value][3]

            elif sheet[f"{indice[IND]}2"].value == 'P6':
                sheet[f"{indice[IND]}3"] = P6[sheet[f"{indice[IND]}1"].value][0]
            elif sheet[f"{indice[IND]}2"].value == 'P15':
                sheet[f"{indice[IND]}3"] = qa[sheet[f"{indice[IND]}1"].value][0]

            elif sheet[f"{indice[IND]}2"].value in ('P13', 'P12', 'P11', 'S501', 'P14'):
                if sheet[f"{indice[IND]}2"].value in ('P11', 'S501'):
                    sheet[f"{indice[IND]}3"] = dsl[sheet[f"{
                        indice[IND]}1"].value][0]
                elif sheet[f"{indice[IND]}2"].value == 'P12':
                    sheet[f"{indice[IND]}3"] = dsl[sheet[f"{
                        indice[IND]}1"].value][1]
                elif sheet[f"{indice[IND]}2"].value == 'P13':
                    sheet[f"{indice[IND]}3"] = dsl[sheet[f"{
                        indice[IND]}1"].value][2]
                else:
                    sheet[f"{indice[IND]}3"] = dsl[sheet[f"{
                        indice[IND]}1"].value][3]

            elif sheet[f"{indice[IND]}2"].value in ('P16', 'P17'):
                if sheet[f"{indice[IND]}2"].value == 'P16':
                    sheet[f"{indice[IND]}3"] = eol[sheet[f"{
                        indice[IND]}1"].value][0]
                else:
                    sheet[f"{indice[IND]}3"] = eol[sheet[f"{
                        indice[IND]}1"].value][1]

            elif sheet[f"{indice[IND]}2"].value in ('PRIN', 'P3', 'P5'):
                if sheet[f"{indice[IND]}2"].value == 'P4':
                    sheet[f"{indice[IND]}3"] = g[sheet[f"{
                        indice[IND]}1"].value][1]
                elif sheet[f"{indice[IND]}2"].value == 'P5':
                    sheet[f"{indice[IND]}3"] = g[sheet[f"{
                        indice[IND]}1"].value][0]
                else:
                    sheet[f"{indice[IND]}3"] = g[sheet[f"{
                        indice[IND]}1"].value][2]
            elif sheet[f"{indice[IND]}2"].value in ('P1', 'P2'):
                if sheet[f"{indice[IND]}2"].value == 'P1':
                    sheet[f"{indice[IND]}3"] = oc[sheet[f"{
                        indice[IND]}1"].value][0]
                else:
                    sheet[f"{indice[IND]}3"] = oc[sheet[f"{
                        indice[IND]}1"].value][1]
            IND += 1
        IND = 4
        for i in range(0, 7):
            sheet[f"B{IND}"] = PO1[sheet[f"A{IND}"].value][0]
            IND += 1
        for i in range(0, 9):
            sheet[f"B{IND}"] = PO2[sheet[f"A{IND}"].value][0]
            IND += 1
    elif i == "Escuros":
        sheet['B4'] = R1['P1'][0]
        sheet['P5'] = R1['P2'][0]
        sheet['B5'] = R2['P1'][0]
        sheet['C5'] = R2['P2'][0]
        sheet['B6'] = R3['P1'][0]
        sheet['C6'] = R3['P2'][0]
        sheet['B7'] = R4['P1'][0]
        sheet['C7'] = R4['P2'][0]
        sheet['B10'] = m_escuros['P1 TA1'][0]
        sheet['C10'] = m_escuros['P2 TA1'][0]
        sheet['B11'] = m_escuros['P1 PO1'][0]
        sheet['C11'] = m_escuros['P2 PO1'][0]
        sheet['B15'] = oc['D6'][0]
        sheet['C15'] = oc['D6'][1]
        sheet['B16'] = oc['D7'][0]
        sheet['C16'] = oc['D7'][1]
        sheet['B17'] = oc['P12'][0]
        sheet['C17'] = oc['P12'][1]
        sheet['B18'] = oc['D9'][0]
        sheet['C18'] = oc['D9'][1]
        sheet['B19'] = oc['D5'][0]
        sheet['C19'] = oc['D5'][1]
        sheet['D19'] = craq['D5'][0]
        sheet['C20'] = oc['D11'][1]
        sheet['B20'] = oc['D11'][0]
        sheet['C21'] = oc['D8'][1]
        sheet['B21'] = oc['D8'][0]
        sheet['C25'] = PO2['P2'][0]
        sheet['B25'] = PO2['P1'][0]
        sheet['B26'] = PO1['P1'][0]
    elif i == "P3":
        sheet['B2'] = R1['P3'][0]
        sheet['B3'] = R2['P3'][0]
        sheet['B4'] = R4['P3'][0]
        sheet['B5'] = R3['P3'][0]

        sheet['B9'] = m_P3['R1'][0]
        sheet['B10'] = m_P3['R2'][0]
        sheet['B11'] = m_P3['R3'][0]
        sheet['B12'] = m_P3['TA1'][0]
        sheet['B15'] = g['D13'][2]
        sheet['B16'] = g['D14'][2]
        sheet['B17'] = g['D12'][2]
        sheet['B18'] = g['D15'][2]
        sheet['B21'] = PO2['P3'][0]
    elif i == 'JET-A':
        sheet['B2'] = R1['P15'][0]
        sheet['B3'] = R2['P15'][0]
        sheet['B7'] = m_jet['R1'][0]
        sheet['B8'] = m_jet['T1'][0]
        sheet['B12'] = qa['D4'][0]
        sheet['B13'] = qa['D30'][0]
        sheet['B14'] = qa['D32'][0]
        sheet['B15'] = qa['D3'][0]

        sheet['B18'] = PO1['P15'][0]
    elif i == "P6":
        sheet['B2'] = R1['P6'][0]
        sheet['B3'] = R2['P6'][0]
        sheet['B4'] = R3['P6'][0]
        sheet['B7'] = m_naf['100%'][0]
        sheet['B8'] = m_naf['95%'][0]
        sheet['B12'] = nft['D4'][0]
        sheet['B13'] = nft['D16'][0]
        sheet['B14'] = nft['D19'][0]
        sheet['B15'] = nft['D3'][0]

        sheet['B16'] = nft['D18-III'][0]
        sheet['B18'] = PO1['P6'][0]
    elif i == "OPASC":
        sheet['C3'] = R5['P3'][0]
        sheet['P5'] = R5['P7'][0]
        sheet['C5'] = R5['P12'][0]
        sheet['C6'] = R5['P11'][0]
        sheet['C7'] = R5['P13'][0]
        sheet['B10'] = m_opasc['R5'][0]
        sheet['B11'] = m_opasc['R5'][1]
        sheet['B12'] = m_opasc['R5'][2]
        sheet['B13'] = m_opasc['R5'][3]
        sheet['B14'] = m_opasc['R5'][4]
        sheet['C10'] = m_opasc['T2'][0]
        sheet['C11'] = m_opasc['T2'][1]
        sheet['C12'] = m_opasc['T2'][2]
        sheet['C13'] = m_opasc['T2'][3]
        sheet['C14'] = m_opasc['T2'][4]
        sheet['P12'] = m_opasc['T3'][0]
        sheet['D11'] = m_opasc['T3'][1]
        sheet['D12'] = m_opasc['T3'][2]
        sheet['D13'] = m_opasc['T3'][3]
        sheet['D14'] = m_opasc['T3'][4]
        sheet['E10'] = m_opasc['T4'][0]
        sheet['E11'] = m_opasc['T4'][1]
        sheet['E12'] = m_opasc['T4'][2]
        sheet['E13'] = m_opasc['T4'][3]
        sheet['E14'] = m_opasc['T4'][4]
    elif i == "Refinarias":
        sheet['B3'] = m_ref['R3'][0]
        sheet['B4'] = m_ref['R3'][1]
        sheet['B5'] = m_ref['R3'][2]
        sheet['B6'] = m_ref['R3'][3]
        sheet['B7'] = m_ref['R3'][4]
        sheet['B8'] = m_ref['R3'][5]
        sheet['B9'] = m_ref['R3'][6]
        sheet['C3'] = m_ref['R1'][0]
        sheet['P5'] = m_ref['R1'][1]
        sheet['C5'] = m_ref['R1'][2]
        sheet['C6'] = m_ref['R1'][3]
        sheet['C7'] = m_ref['R1'][4]
        sheet['C8'] = m_ref['R1'][5]
        sheet['C9'] = m_ref['R1'][6]
        sheet['D3'] = m_ref['R2'][0]
        sheet['D4'] = m_ref['R2'][1]
        sheet['D5'] = m_ref['R2'][2]
        sheet['D6'] = m_ref['R2'][3]
        sheet['D7'] = m_ref['R2'][4]
        sheet['D8'] = m_ref['R2'][5]
        sheet['D9'] = m_ref['R2'][6]
        sheet['E3'] = m_ref['R4'][0]
        sheet['E4'] = m_ref['R4'][1]
        sheet['E5'] = m_ref['R4'][2]
        sheet['E6'] = m_ref['R4'][3]
        sheet['E7'] = m_ref['R4'][4]
        sheet['E8'] = m_ref['R4'][5]
        sheet['E9'] = m_ref['R4'][6]

        sheet['B14'] = emeta_refinarias['R3'][0]
        sheet['B15'] = emeta_refinarias['R3'][1]
        sheet['B16'] = emeta_refinarias['R3'][2]
        sheet['B17'] = emeta_refinarias['R3'][3]
        sheet['B18'] = emeta_refinarias['R3'][4]
        sheet['B19'] = emeta_refinarias['R3'][5]
        sheet['B20'] = emeta_refinarias['R3'][6]

        sheet['B23'] = emeta_refinarias['R3'][7]
        sheet['B24'] = emeta_refinarias['R3'][8]
        sheet['B25'] = emeta_refinarias['R3'][9]

        sheet['C14'] = emeta_refinarias['R1'][0]
        sheet['C15'] = emeta_refinarias['R1'][1]
        sheet['C16'] = emeta_refinarias['R1'][2]
        sheet['C17'] = emeta_refinarias['R1'][3]
        sheet['C18'] = emeta_refinarias['R1'][4]
        sheet['C19'] = emeta_refinarias['R1'][5]
        sheet['C20'] = emeta_refinarias['R1'][6]
        sheet['C23'] = emeta_refinarias['R1'][7]
        sheet['C24'] = emeta_refinarias['R1'][8]
        sheet['C25'] = emeta_refinarias['R1'][9]

        sheet['D14'] = emeta_refinarias['R2'][0]
        sheet['D15'] = emeta_refinarias['R2'][1]
        sheet['D16'] = emeta_refinarias['R2'][2]
        sheet['D17'] = emeta_refinarias['R2'][3]
        sheet['D18'] = emeta_refinarias['R2'][4]
        sheet['D19'] = emeta_refinarias['R2'][5]
        sheet['D20'] = emeta_refinarias['R2'][6]
        sheet['D23'] = emeta_refinarias['R2'][7]
        sheet['D24'] = emeta_refinarias['R2'][8]
        sheet['D25'] = emeta_refinarias['R2'][9]

        sheet['E14'] = emeta_refinarias['R4'][0]
        sheet['E15'] = emeta_refinarias['R4'][1]
        sheet['E16'] = emeta_refinarias['R4'][2]
        sheet['E17'] = emeta_refinarias['R4'][3]
        sheet['E18'] = emeta_refinarias['R4'][4]
        sheet['E19'] = emeta_refinarias['R4'][5]
        sheet['E20'] = emeta_refinarias['R4'][6]
        sheet['E23'] = emeta_refinarias['R4'][7]
        sheet['E24'] = emeta_refinarias['R4'][8]
        sheet['E25'] = emeta_refinarias['R4'][9]

        INDX = 3
        while INDX < 41:
            if sheet[f'G{INDX}'].value == 'R3':
                sheet[f'I{INDX}'] = R3[sheet[f'H{INDX}'].value.strip()][0]
            elif sheet[f'G{INDX}'].value == 'R4':
                sheet[f'I{INDX}'] = R4[sheet[f'H{INDX}'].value.strip()][0]
            elif sheet[f'G{INDX}'].value == 'R2':
                sheet[f'I{INDX}'] = R2[sheet[f'H{INDX}'].value.strip()][0]
            elif sheet[f'G{INDX}'].value == 'R1':
                sheet[f'I{INDX}'] = R1[sheet[f'H{INDX}'].value.strip()][0]
            INDX += 1
    elif i == "Terminais":
        sheet['G4'] = emeta_terminais['T10'][0]
        sheet['B5'] = emeta_terminais['T5'][1]
        sheet['C5'] = emeta_terminais['T6'][1]
        sheet['D5'] = emeta_terminais['T7'][1]
        sheet['E5'] = emeta_terminais['T8'][1]
        sheet['F5'] = emeta_terminais['T9'][1]
        sheet['G5'] = emeta_terminais['T10'][1]
        sheet['H5'] = emeta_terminais['T11'][1]
        sheet['I5'] = emeta_terminais['T12'][1]
        sheet['J5'] = emeta_terminais['TA TA1'][1]
        sheet['K5'] = emeta_terminais['T13'][1]
        sheet['L5'] = emeta_terminais['T14'][1]
        sheet['M5'] = emeta_terminais['T15'][1]

        sheet['G7'] = emeta_terminais['T10'][2]
        sheet['H7'] = emeta_terminais['T11'][2]
        sheet['K7'] = emeta_terminais['T13'][2]

        sheet['G8'] = emeta_terminais['T10'][3]
        sheet['J8'] = emeta_terminais['TA TA1'][3]

        sheet['C9'] = emeta_terminais['T6'][4]
        sheet['D9'] = emeta_terminais['T7'][4]
        sheet['E9'] = emeta_terminais['T8'][4]
        sheet['F9'] = emeta_terminais['T9'][4]
        sheet['G9'] = emeta_terminais['T10'][4]
        sheet['I9'] = emeta_terminais['T12'][4]
        sheet['J9'] = emeta_terminais['TA TA1'][4]
        sheet['K9'] = emeta_terminais['T13'][4]
        sheet['M9'] = emeta_terminais['T15'][4]

        sheet['B10'] = emeta_terminais['T5'][5]
        sheet['C10'] = emeta_terminais['T6'][5]
        sheet['P12'] = emeta_terminais['T7'][5]
        sheet['E10'] = emeta_terminais['T8'][5]
        sheet['F10'] = emeta_terminais['T9'][5]
        sheet['G10'] = emeta_terminais['T10'][5]
        sheet['H10'] = emeta_terminais['T11'][5]
        sheet['I10'] = emeta_terminais['T12'][5]
        sheet['J10'] = emeta_terminais['TA TA1'][5]
        sheet['K10'] = emeta_terminais['T13'][5]
        sheet['L10'] = emeta_terminais['T14'][5]
        sheet['M10'] = emeta_terminais['T15'][5]

        sheet['J13'] = emeta_terminais['TA TA1'][6]
        sheet['J14'] = emeta_terminais['TA TA1'][7]
        sheet['L14'] = emeta_terminais['T14'][7]
        sheet['G14'] = emeta_terminais['T10'][7]
        sheet['M14'] = emeta_terminais['T15'][7]

        sheet['B19'] = PO1["P6"][0]
        sheet['B20'] = PO1["P7"][0]
        sheet['B21'] = PO1["P8"][0]
        sheet['B22'] = PO1["P15"][0]
        sheet['B23'] = PO1["P12"][0]
        sheet['B24'] = PO1["P11"][0]
        sheet['B25'] = PO1["P13"][0]

        sheet['B30'] = PO2['P5'][0]
        sheet['B31'] = PO2['P3'][0]
        sheet['B32'] = PO2['P7'][0]
        sheet['B33'] = PO2['P8'][0]
        sheet['B34'] = PO2['P12'][0]
        sheet['B35'] = PO2['P11'][0]
        sheet['B36'] = PO2['P13'][0]
        sheet['B37'] = PO2['P2'][0]
        sheet['B38'] = PO2['P1'][0]

        sheet['G21'] = m_terminais['T5'][0]
        sheet['H21'] = m_terminais['T6'][0]
        sheet['I21'] = m_terminais['T7'][0]
        sheet['J21'] = m_terminais['T8'][0]
        sheet['K21'] = m_terminais['T9'][0]
        sheet['L21'] = m_terminais['T10'][0]
        sheet['M21'] = m_terminais['T11'][0]
        sheet['N21'] = m_terminais['T12'][0]
        sheet['O21'] = m_terminais['TA TA1'][0]
        sheet['P21'] = m_terminais['T13'][0]
        sheet['Q21'] = m_terminais['T14'][0]
        sheet['R21'] = m_terminais['T15'][0]

        sheet['P23'] = m_terminais['T13'][1]
        sheet['O24'] = m_terminais['TA TA1'][2]

        sheet['H25'] = m_terminais['T6'][3]
        sheet['I25'] = m_terminais['T7'][3]
        sheet['J25'] = m_terminais['T8'][3]
        sheet['K25'] = m_terminais['T9'][3]
        sheet['L25'] = m_terminais['T10'][3]
        sheet['M25'] = m_terminais['T11'][3]
        sheet['N25'] = m_terminais['T12'][3]
        sheet['O25'] = m_terminais['TA TA1'][3]
        sheet['P25'] = m_terminais['T13'][3]
        sheet['R25'] = m_terminais['T15'][3]

        sheet['G26'] = m_terminais['T5'][4]
        sheet['H26'] = m_terminais['T6'][4]
        sheet['I26'] = m_terminais['T7'][4]
        sheet['J26'] = m_terminais['T8'][4]
        sheet['K26'] = m_terminais['T9'][4]
        sheet['L26'] = m_terminais['T10'][4]
        sheet['M26'] = m_terminais['T11'][4]
        sheet['N26'] = m_terminais['T12'][4]
        sheet['O26'] = m_terminais['TA TA1'][4]
        sheet['P26'] = m_terminais['T13'][4]
        sheet['Q26'] = m_terminais['T14'][4]
        sheet['R26'] = m_terminais['T15'][4]

        sheet['O28'] = m_terminais['TA TA1'][5]
        sheet['O29'] = m_terminais['TA TA1'][6]
        sheet['L30'] = m_terminais['T10'][7]
        sheet['O30'] = m_terminais['TA TA1'][7]
    elif i == "Dutos":
        sheet['B4'] = nft['D4'][0]
        sheet['B5'] = glna['D4'][0]
        sheet['B6'] = glna['D4'][1]
        sheet['B7'] = qa['D4'][0]
        sheet['B8'] = dsl['D4'][0]
        sheet['B9'] = dsl['D4'][2]
        sheet['B10'] = eol['TR_D4'][0]
        sheet['B11'] = eol['TR_D4'][1]

        sheet['G4'] = glna['D1'][0]
        sheet['G5'] = dsl['D1'][0]
        sheet['G6'] = dsl['D1'][1]
        sheet['G7'] = dsl['D1'][3]

        sheet['K4'] = nft['D3'][0]
        sheet['K5'] = glna['D3'][0]
        sheet['K6'] = glna['D3'][1]
        sheet['K7'] = qa['D3'][0]
        sheet['K8'] = dsl['D3'][0]
        sheet['K9'] = dsl['D3'][2]
        sheet['K10'] = dsl['D3'][1]

        sheet['N4'] = oc['D9'][1]
        sheet['N5'] = oc['D9'][0]

        sheet['Q4'] = oc['D6'][1]
        sheet['Q5'] = oc['D6'][0]

        sheet['T4'] = glna['D2'][0]
        sheet['T5'] = glna['D2'][1]
        sheet['T6'] = dsl['D2'][0]
        sheet['T7'] = dsl['D2'][3]
        sheet['T8'] = eol['D2'][0]
        sheet['T9'] = eol['D2'][1]

workbook.save(filename=f"bandeira{ANO}{MES}.xlsx")
